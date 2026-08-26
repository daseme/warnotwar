from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook


ARABIC_MONTHS_2025 = {
    "كانون الثاني": 1,
    "شباط": 2,
    "اذار": 3,
    "آذار": 3,
    "نيسان": 4,
    "ايار": 5,
    "أيار": 5,
    "حزيران": 6,
    "تموز": 7,
    "أب": 8,
    "آب": 8,
    "أيلول": 9,
    "تشرين الاول": 10,
    "تشرين الأول": 10,
    "تشرين الثاني": 11,
    "كانون الاول": 12,
    "كانون الأول": 12,
}


@dataclass(frozen=True)
class SomoPeriod:
    period_start: pd.Timestamp
    period_end: pd.Timestamp
    granularity: str
    basrah_barrels: float
    total_barrels: float
    krg_barrels: float | None
    source_file: str
    source_sheet: str

    @property
    def days(self) -> int:
        return (self.period_end - self.period_start).days + 1

    @property
    def total_kbd(self) -> float:
        return self.total_barrels / self.days / 1000.0

    @property
    def basrah_kbd(self) -> float:
        return self.basrah_barrels / self.days / 1000.0


def build_somo_exports(raw_dir: Path) -> pd.DataFrame:
    """Parse official SOMO workbooks into canonical Iraq export periods.

    Canonical policy:
      * 2025 monthly observations come from the annual 2025 workbook.
      * 2026 Jan-Apr come from the individual English monthly sheets.
      * May-Jun 2026 remains one combined official period.
      * December 2025 monthly workbook is used only as a validation cross-check.
    """
    annual_path = raw_dir / "2025_annual.xlsx"
    if not annual_path.exists():
        raise FileNotFoundError(annual_path)

    periods = parse_2025_annual(annual_path)

    monthly_specs = [
        ("2026_january.xlsx", pd.Timestamp("2026-01-01"), pd.Timestamp("2026-01-31"), "month"),
        ("2026_february.xlsx", pd.Timestamp("2026-02-01"), pd.Timestamp("2026-02-28"), "month"),
        ("2026_march.xlsx", pd.Timestamp("2026-03-01"), pd.Timestamp("2026-03-31"), "month"),
        ("2026_april.xlsx", pd.Timestamp("2026-04-01"), pd.Timestamp("2026-04-30"), "month"),
        ("2026_may_june.xlsx", pd.Timestamp("2026-05-01"), pd.Timestamp("2026-06-30"), "multi_month"),
    ]

    for filename, start, end, granularity in monthly_specs:
        path = raw_dir / filename
        if not path.exists():
            raise FileNotFoundError(path)
        periods.append(
            parse_english_summary_workbook(
                path,
                period_start=start,
                period_end=end,
                granularity=granularity,
            )
        )

    december_path = raw_dir / "2025_december.xlsx"
    if december_path.exists():
        december = parse_english_summary_workbook(
            december_path,
            period_start=pd.Timestamp("2025-12-01"),
            period_end=pd.Timestamp("2025-12-31"),
            granularity="month",
        )
        validate_december_against_annual(periods, december)

    frame = periods_to_frame(periods)
    validate_periods(frame)
    return frame


def parse_2025_annual(path: Path) -> list[SomoPeriod]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        periods: list[SomoPeriod] = []

        for row in sheet.iter_rows(min_row=8, max_row=19, values_only=True):
            month_label = normalize_text(row[1])
            month_number = ARABIC_MONTHS_2025.get(month_label)
            if month_number is None:
                raise RuntimeError(
                    f"Unrecognized SOMO 2025 Arabic month: {row[1]!r}"
                )

            start = pd.Timestamp(year=2025, month=month_number, day=1)
            end = start + pd.offsets.MonthEnd(0)

            periods.append(
                SomoPeriod(
                    period_start=start,
                    period_end=end,
                    granularity="month",
                    basrah_barrels=as_number(row[2], "Basrah barrels"),
                    krg_barrels=as_optional_number(row[3]),
                    total_barrels=as_number(row[6], "total barrels"),
                    source_file=path.name,
                    source_sheet=sheet.title,
                )
            )

        return periods
    finally:
        workbook.close()


def parse_english_summary_workbook(
    path: Path,
    *,
    period_start: pd.Timestamp,
    period_end: pd.Timestamp,
    granularity: str,
) -> SomoPeriod:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = choose_english_sheet(workbook.sheetnames)
        ws = workbook[sheet]

        rows = [list(row) for row in ws.iter_rows(values_only=True)]

        header_index = find_header_row(rows)
        header = rows[header_index]
        total_col = find_column(header, "TOTAL")
        basrah_col = find_column(header, "BASRAH CRUDE")
        krg_col = find_column(header, "KRG CRUDE OIL")

        data_row = find_data_row(rows[header_index + 1 :], year=period_start.year)

        return SomoPeriod(
            period_start=period_start,
            period_end=period_end,
            granularity=granularity,
            basrah_barrels=as_number(data_row[basrah_col], "Basrah barrels"),
            krg_barrels=as_optional_number(data_row[krg_col]),
            total_barrels=as_number(data_row[total_col], "total barrels"),
            source_file=path.name,
            source_sheet=sheet,
        )
    finally:
        workbook.close()


def choose_english_sheet(sheetnames: Iterable[str]) -> str:
    for name in sheetnames:
        lowered = name.lower()
        if "انكليزي" in name or "english" in lowered:
            return name
    raise RuntimeError("SOMO workbook has no English publication sheet.")


def find_header_row(rows: list[list[object]]) -> int:
    for index, row in enumerate(rows):
        normalized = {normalize_text(value).upper() for value in row if value is not None}
        if "YEAR" in normalized and "MONTH" in normalized and "TOTAL" in normalized:
            return index
    raise RuntimeError("Could not find SOMO English header row.")


def find_column(header: list[object], label: str) -> int:
    target = normalize_text(label).upper()
    for index, value in enumerate(header):
        if normalize_text(value).upper() == target:
            return index
    raise RuntimeError(f"Could not find SOMO column {label!r}.")


def find_data_row(rows: list[list[object]], year: int) -> list[object]:
    for row in rows:
        for value in row:
            if value == year:
                return row
    raise RuntimeError(f"Could not find SOMO data row for {year}.")


def validate_december_against_annual(
    annual_periods: list[SomoPeriod],
    december: SomoPeriod,
) -> None:
    annual_december = next(
        period
        for period in annual_periods
        if period.period_start == pd.Timestamp("2025-12-01")
    )

    if annual_december.total_barrels != december.total_barrels:
        raise RuntimeError(
            "SOMO December 2025 total differs between annual and monthly workbooks: "
            f"{annual_december.total_barrels} vs {december.total_barrels}"
        )

    if annual_december.basrah_barrels != december.basrah_barrels:
        raise RuntimeError(
            "SOMO December 2025 Basrah barrels differ between annual and monthly workbooks."
        )


def periods_to_frame(periods: Iterable[SomoPeriod]) -> pd.DataFrame:
    rows = []
    for period in periods:
        rows.append(
            {
                "period_start": period.period_start,
                "period_end": period.period_end,
                "granularity": period.granularity,
                "country_code": "IQ",
                "country": "Iraq",
                "basrah_barrels": period.basrah_barrels,
                "krg_barrels": period.krg_barrels,
                "total_barrels": period.total_barrels,
                "days": period.days,
                "basrah_kbd": period.basrah_kbd,
                "total_kbd": period.total_kbd,
                "source": "SOMO",
                "source_file": period.source_file,
                "source_sheet": period.source_sheet,
            }
        )

    return pd.DataFrame(rows).sort_values("period_start").reset_index(drop=True)


def validate_periods(frame: pd.DataFrame) -> None:
    if frame.empty:
        raise RuntimeError("No SOMO export periods parsed.")

    if frame["period_start"].duplicated().any():
        raise RuntimeError("Duplicate SOMO period_start values found.")

    if (frame["period_end"] < frame["period_start"]).any():
        raise RuntimeError("Invalid SOMO period end before start.")

    if (frame[["basrah_barrels", "total_barrels"]] < 0).any().any():
        raise RuntimeError("Negative SOMO export quantities found.")

    if (frame["basrah_barrels"] > frame["total_barrels"]).any():
        raise RuntimeError("SOMO Basrah barrels exceed total barrels.")


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def as_number(value: object, label: str) -> float:
    if value is None or value == "":
        raise RuntimeError(f"Missing SOMO {label}.")
    return float(value)


def as_optional_number(value: object) -> float | None:
    if value is None or value == "":
        return None
    return float(value)
